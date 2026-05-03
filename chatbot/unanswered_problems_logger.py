"""
unanswered_problems_logger.py
-----------------------------------
Backend helper to log unanswered / new problems from Kisan Mitra
into unanswered_problems.xlsx

Call save_problem() from your Flask/Django route when a new
query arrives that is NOT in adv_data.xlsx.

Usage:
    from unanswered_problems_logger import save_problem
    save_problem(query="Meri fasal pe daag aa rahe hain", brief_solution="Fungal - Mancozeb spray")
"""

import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX_PATH = os.path.join(os.path.dirname(__file__), "unanswered_problems.xlsx")

HEADERS = ['Sr No', 'Timestamp', 'User Problem (Query)', 'Brief Solution', 'Category', 'Status']
COL_WIDTHS = [8, 22, 52, 62, 22, 18]

HEADER_FILL  = PatternFill('solid', start_color='2E7D32')
HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=11)
ROW_FILL_ALT = PatternFill('solid', start_color='E8F5E9')
DATA_FONT    = Font(name='Arial', size=10)
BORDER       = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

CATEGORY_MAP = {
    'rog': 'Fasal Rog', 'disease': 'Fasal Rog', 'fungal': 'Fasal Rog', 'dhabb': 'Fasal Rog',
    'keet': 'Keet Niyantran', 'insect': 'Keet Niyantran', 'pest': 'Keet Niyantran',
    'khad': 'Poshan Prabandhan', 'fertil': 'Poshan Prabandhan', 'urea': 'Poshan Prabandhan', 'npk': 'Poshan Prabandhan',
    'sinch': 'Jal Prabandhan', 'water': 'Jal Prabandhan', 'irrigation': 'Jal Prabandhan',
    'mausam': 'Mausam', 'weather': 'Mausam',
}

def detect_category(query: str) -> str:
    q = query.lower()
    for keyword, category in CATEGORY_MAP.items():
        if keyword in q:
            return category
    return 'Anya'


def _ensure_file() -> None:
    """Create xlsx if it doesn't exist."""
    if os.path.exists(XLSX_PATH):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = 'Unanswered Problems'

    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    wb.save(XLSX_PATH)


def save_problem(query: str, brief_solution: str, category: str = None, status: str = 'Pending Review') -> int:
    """
    Append a new unanswered problem row, or update it if it already exists.

    Args:
        query           : The user's original question text
        brief_solution  : One-line solution from adv_data.xlsx or fallback
        category        : Auto-detected if None
        status          : Default 'Pending Review'

    Returns:
        Row number where data was written.
    """
    _ensure_file()

    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    if category is None:
        category = detect_category(query)

    # Search if query already exists
    existing_row = None
    for row in range(2, ws.max_row + 1):
        cell_query = ws.cell(row=row, column=3).value
        if cell_query and cell_query.strip().lower() == query.strip().lower():
            existing_row = row
            break

    timestamp_str = datetime.now().strftime('%Y-%m-%d %H:%M')

    if existing_row:
        # Update existing row
        ws.cell(row=existing_row, column=2, value=timestamp_str)
        ws.cell(row=existing_row, column=4, value=brief_solution)
        ws.cell(row=existing_row, column=6, value=status)
        target_row = existing_row
        print(f"[unanswered_logger] Updated row {target_row} for query: {query[:60]}...")
    else:
        # Find next empty row
        target_row = ws.max_row + 1
        sr_no = target_row - 1  # Header is row 1

        row_data = [
            sr_no,
            timestamp_str,
            query,
            brief_solution,
            category,
            status,
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=target_row, column=col, value=val)
            cell.font   = DATA_FONT
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if target_row % 2 == 0:
                cell.fill = ROW_FILL_ALT

        ws.row_dimensions[target_row].height = 25
        print(f"[unanswered_logger] Saved new row {target_row}: {query[:60]}...")

    # Handle PermissionError when the file is open in Excel
    try:
        wb.save(XLSX_PATH)
    except PermissionError:
        print(f"[unanswered_logger] WARNING: {XLSX_PATH} is open in Excel. Saving to fallback file.")
        fallback_path = XLSX_PATH.replace('.xlsx', '_fallback.xlsx')
        try:
            wb.save(fallback_path)
            print(f"[unanswered_logger] Saved to {fallback_path}")
        except Exception as e:
            print(f"[unanswered_logger] Error saving fallback: {e}")

    return target_row


# -------------------------------------------------------
# CLI test
# -------------------------------------------------------
if __name__ == '__main__':
    print("Testing logger...")
    save_problem(
        query="Meri tamatar ki fasal mein patta mur raha hai",
        brief_solution="Viral disease - copper fungicide + aphid control karein",
        category="Fasal Rog"
    )
    save_problem(
        query="Sarson mein aphid ka attack ho gaya hai",
        brief_solution="Imidacloprid 0.5ml/L water mein spray karein",
    )
    print(f"Done. File: {XLSX_PATH}")